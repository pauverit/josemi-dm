import openpyxl
import json

wb = openpyxl.load_workbook('TARIFA 26.xlsx', data_only=True)

def strip_val(val):
    if val is None: return None
    if isinstance(val, str): return val.strip()
    return val

def clean_text(text):
    if not isinstance(text, str):
        return str(text) if text is not None else ''
    replacements = {
        '\u00c1': 'A', '\u00dd': 'i', '\u00be': 'o', '\u00df': 'a', '\u00b1': 'n',
        '\u2554': 'E', '\u2550': 'I', '\u2534': 'A', '\u00cb': 'O', '\u2551': 'o',
        '\u2524': '!', '\u00b7': 'u', '\u00a2': '1/2',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.strip()

def extract():
    results = {'flexibles': [], 'rigidos': []}
    
    # --- FLEXIBLES ---
    ws = wb['FLEXIBLES']
    curr_cat = None
    curr_subcat = None
    
    for row in range(1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1)
        v1 = strip_val(c1.value)
        v2 = strip_val(ws.cell(row=row, column=2).value)
        v3 = strip_val(ws.cell(row=row, column=3).value)
        
        if v1 is None and v2 is None and v3 is None: continue
        
        # Filter noise
        if v1 and any(x in v1 for x in ['Julian', 'Jose Miguel', 'PRECIO', 'ESPESOR']): continue
        if v2 and any(x in str(v2) for x in ['Julian', 'Jose Miguel']): continue

        # Header detection
        if v1 and c1.font.bold and v3 is None:
            # Main Category if it's all caps or a known major header
            if v1.isupper() or row < 20:
                curr_cat = {'name': clean_text(v1), 'subcategories': []}
                results['flexibles'].append(curr_cat)
                curr_subcat = None
                continue
            # Subcategory
            if curr_cat:
                curr_subcat = {'name': clean_text(v1), 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
                continue
        
        # Product detection
        if v1 and (isinstance(v3, (int, float)) or isinstance(v2, (int, float))):
            prod = {
                'name': clean_text(v1),
                'size': clean_text(v2) if not isinstance(v2, (int, float)) else '',
                'price': v3 if isinstance(v3, (int, float)) else v2
            }
            if not curr_cat:
                curr_cat = {'name': 'General', 'subcategories': []}
                results['flexibles'].append(curr_cat)
            if not curr_subcat:
                curr_subcat = {'name': 'General', 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
            curr_subcat['products'].append(prod)

    # --- RIGIDOS ---
    ws = wb['RIGIDOS']
    curr_cat = None
    curr_subcat = None
    
    for row in range(1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1)
        c2 = ws.cell(row=row, column=2)
        v1 = strip_val(c1.value)
        v2 = strip_val(c2.value)
        v3 = strip_val(ws.cell(row=row, column=3).value)
        
        if v1 is None and v2 is None: continue
        
        # Filter noise
        if v1 and any(x in v1 for x in ['Julian', 'Jose Miguel', 'Espesor', 'rogamos']): continue
        if v2 and any(x in str(v2) for x in ['Julian', 'Jose Miguel', 'PRECIO', 'ESPESOR', 'Disponibles']): continue

        # Section header in col2
        if v2 and c2.font.bold and v1 is None:
            curr_cat = {'name': clean_text(v2), 'subcategories': []}
            results['rigidos'].append(curr_cat)
            curr_subcat = None
            continue
            
        # Subheader in col1
        if v1 and c1.font.bold and (v2 is None or v2 == 'PRECIO'):
            if not curr_cat:
                curr_cat = {'name': 'General', 'subcategories': []}
                results['rigidos'].append(curr_cat)
            curr_subcat = {'name': clean_text(v1), 'products': []}
            curr_cat['subcategories'].append(curr_subcat)
            continue
            
        # Product
        if v1 and isinstance(v2, (int, float)):
            prod = {
                'name': clean_text(v1),
                'price': v2,
                'price_half': v3 if isinstance(v3, (int, float)) else None
            }
            if not curr_subcat:
                if not curr_cat:
                    curr_cat = {'name': 'General', 'subcategories': []}
                    results['rigidos'].append(curr_cat)
                curr_subcat = {'name': 'General', 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
            curr_subcat['products'].append(prod)

    # Remove empties
    results['flexibles'] = [c for c in results['flexibles'] if c['subcategories']]
    results['rigidos'] = [c for c in results['rigidos'] if c['subcategories']]
    
    return results

data = extract()
with open('product_data_v3.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print('Extracted v3 data.')
