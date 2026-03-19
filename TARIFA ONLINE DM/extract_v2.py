import openpyxl
import json

wb = openpyxl.load_workbook('TARIFA 26.xlsx', data_only=True)

def strip_val(val):
    if val is None: return None
    if isinstance(val, str): return val.strip()
    return val

def extract():
    results = {'flexibles': [], 'rigidos': []}
    
    # --- FLEXIBLES ---
    ws = wb['FLEXIBLES']
    curr_cat = None
    curr_subcat = None
    
    for row in range(1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1)
        c2 = ws.cell(row=row, column=2)
        c3 = ws.cell(row=row, column=3)
        
        v1, v2, v3 = strip_val(c1.value), strip_val(c2.value), strip_val(c3.value)
        
        if v1 is None and v2 is None and v3 is None: continue
        
        # Check for category (Bold, no price)
        if v1 and c1.font.bold and v3 is None:
            # If it's a known brand or looks like a subheader, it might be subcat
            # But usually the first few are main categories
            if row < 100 and v1 in ['VINILOS MONOMÉRICOS', 'LAMINADOS MONOMÉRICOS', 'VINILOS POLIMERICOS']:
                curr_cat = {'name': v1, 'subcategories': []}
                results['flexibles'].append(curr_cat)
                curr_subcat = None
                continue
            
            # If we don't have a category yet, this must be it
            if curr_cat is None:
                curr_cat = {'name': v1, 'subcategories': []}
                results['flexibles'].append(curr_cat)
                continue
            
            # Otherwise, it's a subcategory
            curr_subcat = {'name': v1, 'products': []}
            curr_cat['subcategories'].append(curr_subcat)
            continue
            
        # Product row: v1 exists, and v3 (usually price) or v2 exists
        if v1 and (isinstance(v3, (int, float)) or isinstance(v2, (int, float))):
            prod = {
                'name': v1,
                'size': v2 if not isinstance(v2, (int, float)) else '',
                'price': v3 if isinstance(v3, (int, float)) else v2
            }
            if curr_subcat:
                curr_subcat['products'].append(prod)
            elif curr_cat:
                # Add to a default "General" subcat if none exists
                if not curr_cat['subcategories'] or curr_cat['subcategories'][-1]['name'] != 'General':
                    curr_cat['subcategories'].append({'name': 'General', 'products': []})
                curr_cat['subcategories'][-1]['products'].append(prod)
                
    # --- RIGIDOS ---
    ws = wb['RIGIDOS']
    curr_cat = None
    curr_subcat = None
    
    for row in range(1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1)
        c2 = ws.cell(row=row, column=2)
        c1_val = strip_val(c1.value)
        c2_val = strip_val(c2.value)
        
        if c1_val is None and c2_val is None: continue
        
        # Rigidos uses col2 for main category often (bold)
        if c2_val and c2.font.bold and c1_val is None:
            curr_cat = {'name': c2_val, 'subcategories': []}
            results['rigidos'].append(curr_cat)
            curr_subcat = None
            continue
            
        # Subcategory: col1 is bold, col2 might say 'PRECIO'
        if c1_val and c1.font.bold and (c2_val is None or c2_val == 'PRECIO'):
            if curr_cat is None:
                curr_cat = {'name': c1_val, 'subcategories': []}
                results['rigidos'].append(curr_cat)
            curr_subcat = {'name': c1_val, 'products': []}
            curr_cat['subcategories'].append(curr_subcat)
            continue
            
        # Product: col1 is name (thickness), col2 is price
        if c1_val and isinstance(c2_val, (int, float)):
            prod = {
                'name': c1_val,
                'price': c2_val,
                'price_half': strip_val(ws.cell(row=row, column=3).value)
            }
            if curr_subcat:
                curr_subcat['products'].append(prod)
            elif curr_cat:
                if not curr_cat['subcategories'] or curr_cat['subcategories'][-1]['name'] != 'General':
                    curr_cat['subcategories'].append({'name': 'General', 'products': []})
                curr_cat['subcategories'][-1]['products'].append(prod)
                
    return results

data = extract()
with open('product_data_v2.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print('Extracted v2 data.')
