import openpyxl
import json
import os

wb = openpyxl.load_workbook('TARIFA 26.xlsx', data_only=True)

def strip_val(val):
    if val is None: return None
    if isinstance(val, str): return val.strip()
    return val

def clean_text(text):
    if not isinstance(text, str):
        return str(text) if text is not None else ''
    replacements = {
        '\u00c1': 'A', '\u00dd': 'i', '\u00be': 'o', '\u00df': 'a', '\u00b1': 'n',
        '\u2554': 'E', '\u2550': 'I', '\u2534': 'A', '\u00cb': 'O', '\u2551': 'o',
        '\u2524': '!', '\u00b7': 'u', '\u00a2': '1/2',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.strip()

# FLEXIBLES WHITELIST
FLEX_CATEGORY_WHITELIST = {
    'VINILOS MONOMÉRICOS': 'VINILO MONO',
    'LAMINADOS MONOMÉRICOS': 'LAMINADO MONO',
    'VINILOS POLIMÉRICOS': 'VINILO POLI',
    'VINILOS POLIMERICOS': 'VINILO POLI',
    'LAMINADOS POLIMÉRICOS': 'LAMINADO POLI',
    'LAMINADOS POLIMERICOS': 'LAMINADO POLI',
    'VINILO FUNDICION': 'VINILO CAST',
    'CAST 50': 'VINILO CAST',
    'LAMINADO FUNDICIÓN': 'LAMINADO CAST',
    'LAMINADO RITRAMA CAST': 'LAMINADO CAST',
    'LONAS FUNDIDAS': 'LONAS',
    'LONAS LAMINADAS': 'LONAS',
    'LONA MESH': 'LONAS',
    'PAPELES': 'PAPELES',
    'JUNKERS & MUELLERS': 'TEXTILES',
    'MEDIATEX': 'TEXTILES',
    'ESPECIALIDADES EN CANVAS': 'CANVAS ARPLEN',
    'VINILO WINDOW': 'VINILO WINDOW',
    'POLIPROPILENO': 'POLIPROPILENO',
    'ESPECIAL ROLL UPS': 'ESPECIAL ROLL UPS',
    'PROTECCIÓN SOLAR': 'PROTECCIÓN SOLAR',
    'TINTAS': 'TINTAS',
    'HP - BMG': 'HP - BMG'
}

def extract_flexibles():
    ws = wb['FLEXIBLES']
    categories = []
    
    # 1. ADD FEDRIGONI SECTION (FROM PDF DATA)
    fed_cat = {
        'name': 'VINILO M96 FEDRIGONI',
        'subcategories': [{
            'name': 'Ri-Jet M96 ESPECIAL LATEX MONOMERICO',
            'products': [
                {'name': 'Ri-Jet M96 Clear Gloss Perm PE', 'size': 'PVP', 'price': 1.37},
                {'name': 'Ri-Jet M96 White Gloss Perm Grey Airflow', 'size': 'PVP', 'price': 1.60},
                {'name': 'Ri-Jet M96 White Gloss Perm Grey PE', 'size': 'PVP', 'price': 1.37},
                {'name': 'Ri-Jet M96 White Gloss Perm PE', 'size': 'PVP', 'price': 1.37},
                {'name': 'Ri-Jet M96 White Gloss Rem Grey Airflow', 'size': 'PVP', 'price': 1.69},
                {'name': 'Ri-Jet M96 White Gloss Rem Grey PE', 'size': 'PVP', 'price': 1.46},
                {'name': 'Ri-Jet M96 White Gloss Rem PE', 'size': 'PVP', 'price': 1.46},
                {'name': 'Ri-Jet M96 White Matt Perm Grey PE', 'size': 'PVP', 'price': 1.40},
                {'name': 'Ri-Jet M96 White Matt Perm PE', 'size': 'PVP', 'price': 1.40},
                {'name': 'Ri-Jet M96 White Matt Rem Grey PE', 'size': 'PVP', 'price': 1.49},
                {'name': 'Ri-Jet M96 White Matt Rem PE', 'size': 'PVP', 'price': 1.49}
            ]
        }]
    }
    categories.append(fed_cat)

    curr_cat = None
    curr_subcat = None

    for row in range(1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1)
        v1 = strip_val(c1.value)
        v2 = strip_val(ws.cell(row=row, column=2).value)
        v3 = strip_val(ws.cell(row=row, column=3).value)
        
        if v1 is None: continue
        
        # Header Detection
        if c1.font.bold and v3 is None:
            v1_up = v1.upper()
            mapped_name = None
            for key, val in FLEX_CATEGORY_WHITELIST.items():
                if key in v1_up:
                    mapped_name = val
                    break
            
            if mapped_name:
                existing = next((c for c in categories if c['name'] == mapped_name), None)
                if not existing:
                    curr_cat = {'name': mapped_name, 'subcategories': []}
                    categories.append(curr_cat)
                else:
                    curr_cat = existing
                
                curr_subcat = {'name': clean_text(v1), 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
                continue

        if v1 and (isinstance(v3, (int, float)) or isinstance(v2, (int, float))):
            if any(x in v1 for x in ['Julian', 'Jose Miguel', 'PRECIO', 'ESPESOR', 'rogamos']): continue
            if not curr_cat: continue 

            prod = {
                'name': clean_text(v1),
                'size': clean_text(v2) if not isinstance(v2, (int, float)) else '',
                'price': v3 if isinstance(v3, (int, float)) else v2
            }
            if not curr_subcat:
                curr_subcat = {'name': 'General', 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
            curr_subcat['products'].append(prod)

    for cat in categories:
        cat['subcategories'] = [sc for sc in cat['subcategories'] if sc['products']]
    return [c for c in categories if c['subcategories']]

def extract_rigidos():
    ws = wb['RIGIDOS']
    categories = []
    
    # 1. ADD PEGASUS (FROM PDF)
    pegasus_cat = {
        'name': 'Pegasus (Panelplast)',
        'subcategories': [{
            'name': 'PANELES PEGASUS - PRECIOS ACTUALIZADOS',
            'products': [
                {'name': '3050x2030 10mm (Film PVC Blanco)', 'price': 70.15, 'price_half': 59.42},
                {'name': '3050x1530 10mm (Film PVC Blanco)', 'price': 53.59, 'price_half': 44.75},
                {'name': '3050x2030 19mm (Film PVC Blanco)', 'price': 95.00},
                {'name': '3050x2030 30mm (Film PVC Blanco)', 'price': 162.00}
            ]
        }]
    }
    categories.append(pegasus_cat)

    # 2. ADD FOREX LITE (FROM PDF)
    forex_cat = {
        'name': 'Forex Lite',
        'subcategories': [{
            'name': 'FOREX® LITE - PVC ESPUMADO',
            'products': [
                {'name': '3mm 2030x3050', 'price': 27.75, 'price_half': 23.79},
                {'name': '5mm 2030x3050', 'price': 46.97, 'price_half': 40.26},
                {'name': '10mm 2030x3050', 'price': 94.48, 'price_half': 81.00}
            ]
        }]
    }
    categories.append(forex_cat)

    curr_cat = None
    curr_subcat = None
    NAMES = ['Julian', 'Jose Miguel']

    for row in range(1, ws.max_row + 1):
        if 190 <= row <= 240: continue # Skip Pegasus/Pluma rows (we use PDF data or specific logic)
        row_vals = [strip_val(ws.cell(row=row, column=c).value) for c in range(1, 10)]
        row_fonts = [ws.cell(row=row, column=c).font.bold for c in range(1, 10)]
        v1, v2, v3 = row_vals[0], row_vals[1], row_vals[2]
        
        # SKIP if it's the areas we manually added or overlapping
        if isinstance(v2, str) and ('Pegasus' in v2 or 'Carton Pluma' in v2):
            pass # We catch them below but avoid mixing

        if not any(row_vals): continue
        
        # Category headers in col2 or col3
        if (v2 and row_fonts[1] and not v1) or (v3 == 'Carton Pluma' and not v1 and not v2):
            cat_name = v2 if v2 else v3
            if any(n in cat_name for n in NAMES): continue
            if 'PEGASUS' in cat_name.upper(): continue # Skip, we have PDF data
            
            curr_cat = {'name': clean_text(cat_name), 'subcategories': []}
            categories.append(curr_cat)
            curr_subcat = None
            continue
            
        # Subcategory
        if v1 and row_fonts[0] and (not v2 or v2 in ['PRECIO', 'MEDIDAS', 'ESPESOR']):
            if v1 in NAMES: continue
            if not curr_cat:
                curr_cat = {'name': 'General', 'subcategories': []}
                categories.append(curr_cat)
            curr_subcat = {'name': clean_text(v1), 'products': []}
            curr_cat['subcategories'].append(curr_subcat)
            continue
            
        # Product
        price = None
        price_half = None
        desc_parts = []
        for i, val in enumerate(row_vals):
            if isinstance(val, (int, float)) and price is None:
                price = val
                if i+1 < len(row_vals) and isinstance(row_vals[i+1], (int, float)):
                    price_half = row_vals[i+1]
                break
            elif isinstance(val, str) and price is None:
                if val.upper() not in ['PRECIO', 'ESPESOR', 'MEDIDAS', 'COLORES', '€ PLACA', 'PLANCHAS CAJA', 'GEOFOAM']:
                    desc_parts.append(val)
        
        if price is not None:
            full_name = " ".join(desc_parts)
            if any(n in full_name for n in NAMES): continue
            
            prod = {
                'name': clean_text(full_name),
                'price': price,
                'price_half': price_half
            }
            if not curr_subcat:
                if not curr_cat:
                    curr_cat = {'name': 'General', 'subcategories': []}
                    categories.append(curr_cat)
                curr_subcat = {'name': 'General', 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
            curr_subcat['products'].append(prod)

    for cat in categories:
        cat['subcategories'] = [sc for sc in cat['subcategories'] if sc['products']]
    return [c for c in categories if c['subcategories']]

def reorganize_tintas(tintas_list):
    all_prods = []
    for cat in tintas_list:
        if 'subcategories' not in cat: continue
        for sc in cat['subcategories']:
            for p in sc['products']:
                all_prods.append(p)
    
    buckets = {
        'Tintas 400ml': [],
        'Tintas 775ml': [],
        'Tintas 1L': [],
        'Tintas 3L': [],
        'Tintas 5L': [],
        'Tintas 10L': [],
        'Cabezales': [],
        'Kit Mantenimiento': [],
        'DTF': [],
        'DTF UV': [],
        'UV': [],
        'Otros Consumibles': []
    }
    
    for p in all_prods:
        name = p['name'].upper()
        size = str(p.get('size', '')).upper()
        
        if 'DTF UV' in name or 'DTF UV' in size: buckets['DTF UV'].append(p)
        elif 'DTF' in name or 'DTF' in size: buckets['DTF'].append(p)
        elif 'CABEZAL' in name: buckets['Cabezales'].append(p)
        elif any(x in name for x in ['MANTENIMIENTO', 'MAINTENANCE', 'CLEANING', 'LIMPIEZA', 'WASTE']): buckets['Kit Mantenimiento'].append(p)
        elif ' UV' in name or ' UV' in size: buckets['UV'].append(p)
        elif '400ML' in name or '400ML' in size or '400 ML' in name: buckets['Tintas 400ml'].append(p)
        elif '775ML' in name or '775ML' in size or '775 ML' in name: buckets['Tintas 775ml'].append(p)
        elif '1L' in name or '1L' in size or '1000ML' in name: buckets['Tintas 1L'].append(p)
        elif '3L' in name or '3L' in size or '3000ML' in name: buckets['Tintas 3L'].append(p)
        elif '5L' in name or '5L' in size or '5000ML' in name: buckets['Tintas 5L'].append(p)
        elif '10L' in name or '10L' in size or '10000ML' in name: buckets['Tintas 10L'].append(p)
        else: buckets['Otros Consumibles'].append(p)
        
    new_cats = []
    for b_name, b_prods in buckets.items():
        if b_prods:
            # Each bucket becomes a top-level category for the sidebar
            new_cats.append({
                'name': b_name, 
                'subcategories': [{'name': 'General', 'products': b_prods}]
            })
    
    return new_cats

def extract_main():
    flex_all = extract_flexibles()
    tinta_names = ['TINTAS', 'HP - BMG']
    flexibles = [c for c in flex_all if c['name'] not in tinta_names]
    tintas_raw = [c for c in flex_all if c['name'] in tinta_names]
    tintas = reorganize_tintas(tintas_raw)
    
    rigidos = extract_rigidos()
    
    return {
        'flexibles': flexibles,
        'rigidos': rigidos,
        'tintas': tintas
    }

data = extract_main()

with open('product_data_v6.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print('Extracted and REORGANIZED v6 data.')
