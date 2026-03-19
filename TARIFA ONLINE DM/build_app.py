import openpyxl
import json
import re

wb = openpyxl.load_workbook('TARIFA 26.xlsx', data_only=True)

def clean_text(text):
    if not isinstance(text, str):
        return str(text) if text is not None else ''
    replacements = {
        '\u00c1': 'µ', '\u00dd': 'í', '\u00be': 'ó', '\u00df': 'á', '\u00b1': 'ñ',
        '\u2554': 'É', '\u2550': 'Í', '\u2534': 'Á', '\u00cb': 'Ó', '\u2551': 'º',
        '\u2524': '¡', '\u00b7': 'ú', '\u00a2': '½',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.strip()

def extract_all_data():
    result = {"flexibles": [], "rigidos": []}
    
    # ============ FLEXIBLES ============
    ws = wb['FLEXIBLES']
    merged_rows = set()
    for mc in ws.merged_cells.ranges:
        merged_rows.add(mc.min_row)
    
    current_cat = None
    current_subcat = None
    current_brand = None
    brands = ['TMK', 'RITRAMA', 'ATP', 'HP']
    
    for row in range(1, ws.max_row + 1):
        vals = {}
        for col in range(1, 25):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                vals[col] = v
        if not vals:
            continue
        
        c1 = clean_text(vals.get(1, ''))
        c2 = clean_text(vals.get(2, ''))
        c3 = vals.get(3, None)
        
        if not c1 and not c2:
            continue
        if c1 == '.':
            continue
        
        # Category (merged)
        if row in merged_rows and c1:
            cat_name = c1
            # Merge duplicate vinilos monomericos
            if current_cat and current_cat['name'] == cat_name:
                continue
            current_cat = {'name': cat_name, 'products': []}
            current_subcat = None
            current_brand = None
            result['flexibles'].append(current_cat)
            continue
        
        if not current_cat:
            continue
        
        # Brand
        if c1 in brands and not c2 and not isinstance(c3, (int, float)):
            current_brand = c1
            continue
        
        # Spec/subheader (no price in col3)
        if c1 and not c2 and not isinstance(c3, (int, float)):
            # Skip descriptions and notes
            skip_keywords = ['VALIDO PARA', 'Medidas', 'Medida', 'PARA ', 'B1 ', 'M1 ', 
                           'CORTE', 'TEXTIL', 'TELA ', 'POLICANVAS', 'ROLL UP', 'Wall ',
                           'IMPERMEAB', '100%PES', '100% PES', 'LATEX UV', 'LATEX Y',
                           'Y RESISTENCIA', 'NO SE ARRUGA', 'Description', 'BANNER AND',
                           'CANVAS MATERIALS', 'FABRICS', 'FILMS', 'FILM ART', 'PAPERS',
                           'SELF-ADH', 'WALL COV', 'POLIESTER', 'FLEXIBLE', 'SUAVE']
            if any(c1.startswith(kw) or c1.startswith(' ' + kw) for kw in skip_keywords):
                continue
            if any(kw in c1 for kw in ['100%PES', '100% PES', 'LATEX UV SOLVENTE', 'LATEX', 'SOLVENTE']):
                if len(c1) < 30:
                    continue
            current_subcat = c1
            continue
        
        # Product with price in col3
        if c1 and isinstance(c3, (int, float)):
            prod = {
                'name': c1,
                'size': c2,
                'price': round(float(c3), 2)
            }
            if current_brand:
                prod['brand'] = current_brand
            if current_subcat:
                prod['spec'] = current_subcat
            current_cat['products'].append(prod)
            continue
    
    # ============ RIGIDOS ============
    ws = wb['RIGIDOS']
    current_cat = None
    current_spec = None
    
    for row in range(1, ws.max_row + 1):
        vals = {}
        for col in range(1, 27):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                vals[col] = v
        if not vals:
            continue
        
        c1 = clean_text(vals.get(1, ''))
        c2 = vals.get(2, '')
        c3 = vals.get(3, '')
        c4 = vals.get(4, '')
        c5 = vals.get(5, '')
        
        c2_str = clean_text(str(c2)) if c2 else ''
        c3_str = clean_text(str(c3)) if c3 else ''
        
        # Skip info rows
        if 'Julian' in c1 or 'Julian' in c2_str or 'Jose Miguel' in c1 or 'Jose Miguel' in c2_str:
            continue
        if 'Disponibles' in c3_str or 'rogamos' in c1 or 'Panel sándwich' in c1 or 'recubierto' in c1:
            continue
        if c1 == 'DISPONIBILIDAD EN COLOR NEGRO. CONSULTAR' or c1.startswith('DISPONIBILIDAD'):
            continue
        if c1 == 'Espesor' or c1 == 'MM' or c1 == 'MEDIDAS' or c2_str == 'PRECIO' or c2_str == 'ESPESOR' or c2_str == 'MEDIDA':
            continue
        if 'GEOFOAM' in c1:
            current_spec = c1
            continue
        
        # Major section headers
        if 'PVC BLANCO' in c1 or 'PVC NEGRO' in c1 or 'PVC Blanco/Nucleo' in c1 or 'PVC Efecto Madera' in c1:
            current_cat = {'name': c1, 'products': []}
            current_spec = None
            result['rigidos'].append(current_cat)
            continue
        
        if not c1 and c2_str:
            sections = ['Planchas de PVC', 'Policarbonato Celular', 'Metacrilato', 'Composite', 'Panelplast']
            for s in sections:
                if s in c2_str:
                    current_cat = {'name': c2_str, 'products': []}
                    current_spec = None
                    result['rigidos'].append(current_cat)
                    break
            continue
        
        if 'Carton Pluma' in c3_str or 'Cartón Pluma' in c3_str:
            current_cat = {'name': 'Cartón Pluma', 'products': []}
            current_spec = None
            result['rigidos'].append(current_cat)
            continue
        
        if 'POLICARBONATO COMPUESTO' in c1:
            current_cat = {'name': c1, 'products': []}
            current_spec = None
            result['rigidos'].append(current_cat)
            continue
        
        if 'POLIPROPULENO ALVEOLAR' in c1 or 'NIDO ABEJA' in c1:
            current_cat = {'name': c1, 'products': []}
            current_spec = None
            result['rigidos'].append(current_cat)
            continue
        
        if 'PERFIL' in c1:
            current_spec = c1
            continue
        
        if c1 in ['COLADA', 'EXTRUSION']:
            current_spec = c1
            continue
        
        # Note rows
        if isinstance(c1, str) and ('Blanco Roto' in c1 or 'Blanco y Hielo' in c1 or 'COLORES:' in c1):
            if 'COLORES:' in c1:
                current_spec = c1
            continue
        
        if not current_cat:
            continue
        
        # Products: price in col2
        if c1 and isinstance(c2, (int, float)):
            prod = {'name': c1, 'price': round(float(c2), 2)}
            if isinstance(c3, (int, float)):
                prod['price_half'] = round(float(c3), 2)
            if current_spec:
                prod['spec'] = current_spec
            current_cat['products'].append(prod)
            continue
        
        # Products: price in col5 (composite)
        if c1 and isinstance(c5, (int, float)):
            prod = {
                'name': c1,
                'price': round(float(c5), 2)
            }
            if c2_str:
                prod['thickness'] = c2_str
            if isinstance(c3, str) and c3:
                prod['color'] = c3_str
            if current_spec:
                prod['spec'] = current_spec
            current_cat['products'].append(prod)
            continue
        
        # Products: price in col4 (carton pluma)
        if c1 and isinstance(c4, (int, float)):
            prod = {
                'name': c1,
                'price': round(float(c4), 2)
            }
            if c2_str:
                prod['size'] = c2_str
            if isinstance(c3, (int, float)):
                prod['units_box'] = int(c3)
            if current_spec:
                prod['spec'] = current_spec
            current_cat['products'].append(prod)
            continue
    
    return result

data = extract_all_data()

# Remove empty categories
data['flexibles'] = [c for c in data['flexibles'] if c['products']]
data['rigidos'] = [c for c in data['rigidos'] if c['products']]

print("FLEXIBLES:")
for cat in data['flexibles']:
    print(f"  {cat['name']}: {len(cat['products'])} products")
print(f"\nRIGIDOS:")
for cat in data['rigidos']:
    print(f"  {cat['name']}: {len(cat['products'])} products")

total_f = sum(len(c['products']) for c in data['flexibles'])
total_r = sum(len(c['products']) for c in data['rigidos'])
print(f"\nTotal: {total_f} flexibles + {total_r} rigidos = {total_f + total_r}")

with open('product_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("Saved to product_data.json")
