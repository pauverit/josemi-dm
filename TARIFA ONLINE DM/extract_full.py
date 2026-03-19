import openpyxl
import json

wb = openpyxl.load_workbook('TARIFA 26.xlsx', data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*60}")
    
    # Find category rows (merged cells spanning A:C usually)
    merged_rows = set()
    for mc in ws.merged_cells.ranges:
        merged_rows.add(mc.min_row)
    
    # Print ALL rows with data
    for row in range(1, ws.max_row + 1):
        vals = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                vals[col] = v
        if vals:
            is_merged = row in merged_rows
            prefix = ">>CATEGORY<<" if is_merged else ""
            print(f"Row {row} {prefix}: {vals}")
