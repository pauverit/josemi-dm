import openpyxl
import json

wb = openpyxl.load_workbook('TARIFA 26.xlsx', data_only=True)

def strip_val(val):
    if val is None: return None
    if isinstance(val, str): return val.strip()
    return val

def clean_text(text):
    if not isinstance(text, str):
        return str(text) if text is not None else ''
    replacements = {
        '\u00c1': 'A', '\u00dd': 'i', '\u00be': 'o', '\u00df': 'a', '\u00b1': 'n',
        '\u2554': 'E', '\u2550': 'I', '\u2534': 'A', '\u00cb': 'O', '\u2551': 'o',
        '\u2524': '!', '\u00b7': 'u', '\u00a2': '1/2',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.strip()

# Map Excel headers to User's requested categories
CATEGORY_MAP = {
    'VINILOS MONOMÉRICOS': 'VINILO MONO',
    'LAMINADOS MONOMÉRICOS': 'LAMINADO MONO',
    'VINILOS POLIMÉRICOS': 'VINILO POLI',
    'VINILOS POLIMERICOS': 'VINILO POLI',
    'LAMINADOS POLIMÉRICOS': 'LAMINADO POLI',
    'LAMINADOS POLIMERICOS': 'LAMINADO POLI',
    'LAMINADO FUNDICIÓN': 'LAMINADO CAST',
    'LAMINADO RITRAMA CAST': 'LAMINADO CAST',
    'CAST 50': 'VINILO CAST',
    'LONAS FUNDIDAS': 'LONAS',
    'LONAS LAMINADAS': 'LONAS',
    'LONA MESH': 'LONAS',
    'PAPELES': 'PAPELES',
    'JUNKERS & MUELLERS': 'TEXTILES',
    'MEDIATEX': 'TEXTILES',
    'ESPECIALIDADES EN CANVAS': 'CANVAS ARPLEN',
    'VINILO WINDOW': 'VINILO WINDOW',
    'POLIPROPILENO': 'POLIPROPILENO',
    'ESPECIAL ROLL UPS': 'ESPECIAL ROLL UPS',
    'PROTECCIÓN SOLAR': 'PROTECCIÓN SOLAR',
    'LAMINAS PROTECCION SOLAR': 'PROTECCIÓN SOLAR',
    'TINTAS': 'TINTAS',
    'HP - BMG': 'HP - BMG'
}

def extract_flexibles():
    ws = wb['FLEXIBLES']
    categories = []
    curr_cat = None
    curr_subcat = None

    for row in range(1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1)
        v1 = strip_val(c1.value)
        v2 = strip_val(ws.cell(row=row, column=2).value)
        v3 = strip_val(ws.cell(row=row, column=3).value)
        
        if v1 is None: continue
        
        # Header Detection (Bold)
        if c1.font.bold and v3 is None:
            v1_up = v1.upper()
            mapped_name = None
            for key, val in CATEGORY_MAP.items():
                if key in v1_up:
                    mapped_name = val
                    break
            
            if mapped_name:
                # If we already have this category, just point to it
                existing = next((c for c in categories if c['name'] == mapped_name), None)
                if existing:
                    curr_cat = existing
                else:
                    curr_cat = {'name': mapped_name, 'subcategories': []}
                    categories.append(curr_cat)
                
                # Treat the actual header text as a subcategory
                curr_subcat = {'name': clean_text(v1), 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
                continue
            
            # If bold but not a major category, it might be a subcategory
            if curr_cat:
                curr_subcat = {'name': clean_text(v1), 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
                continue

        # Product Detection
        if v1 and (isinstance(v3, (int, float)) or isinstance(v2, (int, float))):
            if any(x in v1 for x in ['Julian', 'Jose Miguel', 'PRECIO', 'ESPESOR']): continue
            
            # If no cat/subcat yet, put in Misc (shouldn't happen with strict headers)
            if not curr_cat: continue 

            prod = {
                'name': clean_text(v1),
                'size': clean_text(v2) if not isinstance(v2, (int, float)) else '',
                'price': v3 if isinstance(v3, (int, float)) else v2
            }
            if not curr_subcat:
                curr_subcat = {'name': 'General', 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
            curr_subcat['products'].append(prod)

    # Post-processing: Merge subcategories with same name or consolidate empty ones
    for cat in categories:
        new_subcats = []
        for sc in cat['subcategories']:
            if not sc['products']: continue
            new_subcats.append(sc)
        cat['subcategories'] = new_subcats

    return [c for c in categories if c['subcategories']]

def extract_rigidos():
    ws = wb['RIGIDOS']
    categories = []
    curr_cat = None
    curr_subcat = None
    
    for row in range(1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1)
        c2 = ws.cell(row=row, column=2)
        v1 = strip_val(c1.value)
        v2 = strip_val(c2.value)
        v3 = strip_val(ws.cell(row=row, column=3).value)
        
        if v1 is None and v2 is None: continue
        
        # Section header in col2
        if v2 and c2.font.bold and v1 is None:
            if any(x in v2 for x in ['Julian', 'Jose Miguel']): continue
            curr_cat = {'name': clean_text(v2), 'subcategories': []}
            categories.append(curr_cat)
            curr_subcat = None
            continue
            
        # Subheader in col1
        if v1 and c1.font.bold and (v2 is None or v2 == 'PRECIO'):
            if any(x in v1 for x in ['Julian', 'Jose Miguel']): continue
            if not curr_cat:
                curr_cat = {'name': 'General', 'subcategories': []}
                categories.append(curr_cat)
            curr_subcat = {'name': clean_text(v1), 'products': []}
            curr_cat['subcategories'].append(curr_subcat)
            continue
            
        # Product
        if v1 and isinstance(v2, (int, float)):
            if any(x in v1 for x in ['Julian', 'Jose Miguel', 'Espesor', 'rogamos']): continue
            prod = {
                'name': clean_text(v1),
                'price': v2,
                'price_half': v3 if isinstance(v3, (int, float)) else None
            }
            if not curr_subcat:
                curr_subcat = {'name': 'General', 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
            curr_subcat['products'].append(prod)

    return [c for c in categories if c['subcategories']]

data = {
    'flexibles': extract_flexibles(),
    'rigidos': extract_rigidos()
}

with open('product_data_v4.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print('Extracted v4 data.')
