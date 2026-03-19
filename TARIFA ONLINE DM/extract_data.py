import openpyxl
import json

wb = openpyxl.load_workbook('TARIFA 26.xlsx', data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)")
    print(f"{'='*60}")
    
    # Print merged cells
    print(f"\nMerged cells: {list(ws.merged_cells.ranges)[:20]}")
    
    # Print first 50 non-empty rows
    count = 0
    for row in range(1, min(ws.max_row+1, 100)):
        vals = {}
        for col in range(1, ws.max_column+1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                vals[col] = v
        if vals:
            print(f"Row {row}: {vals}")
            count += 1
