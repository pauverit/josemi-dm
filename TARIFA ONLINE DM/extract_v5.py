import openpyxl
import json

wb = openpyxl.load_workbook('TARIFA 26.xlsx', data_only=True)

def strip_val(val):
    if val is None: return None
    if isinstance(val, str): return val.strip()
    return val

def clean_text(text):
    if not isinstance(text, str):
        return str(text) if text is not None else ''
    replacements = {
        '\u00c1': 'A', '\u00dd': 'i', '\u00be': 'o', '\u00df': 'a', '\u00b1': 'n',
        '\u2554': 'E', '\u2550': 'I', '\u2534': 'A', '\u00cb': 'O', '\u2551': 'o',
        '\u2524': '!', '\u00b7': 'u', '\u00a2': '1/2',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.strip()

# Strictly follow the user's requested list for FLEXIBLES
FLEX_CATEGORY_WHITELIST = {
    'VINILOS MONOMÉRICOS': 'VINILO MONO',
    'LAMINADOS MONOMÉRICOS': 'LAMINADO MONO',
    'VINILOS POLIMÉRICOS': 'VINILO POLI',
    'VINILOS POLIMERICOS': 'VINILO POLI',
    'LAMINADOS POLIMÉRICOS': 'LAMINADO POLI',
    'LAMINADOS POLIMERICOS': 'LAMINADO POLI',
    'VINILO FUNDICION': 'VINILO CAST',
    'CAST 50': 'VINILO CAST',
    'LAMINADO FUNDICIÓN': 'LAMINADO CAST',
    'LAMINADO RITRAMA CAST': 'LAMINADO CAST',
    'LONAS FUNDIDAS': 'LONAS',
    'LONAS LAMINADAS': 'LONAS',
    'LONA MESH': 'LONAS',
    'PAPELES': 'PAPELES',
    'JUNKERS & MUELLERS': 'TEXTILES',
    'MEDIATEX': 'TEXTILES',
    'ESPECIALIDADES EN CANVAS': 'CANVAS ARPLEN',
    'VINILO WINDOW': 'VINILO WINDOW',
    'POLIPROPILENO': 'POLIPROPILENO',
    'ESPECIAL ROLL UPS': 'ESPECIAL ROLL UPS',
    'PROTECCIÓN SOLAR': 'PROTECCIÓN SOLAR',
    'TINTAS': 'TINTAS',
    'HP - BMG': 'HP - BMG'
}

def extract_flexibles():
    ws = wb['FLEXIBLES']
    categories = []
    curr_cat = None
    curr_subcat = None

    for row in range(1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1)
        v1 = strip_val(c1.value)
        v2 = strip_val(ws.cell(row=row, column=2).value)
        v3 = strip_val(ws.cell(row=row, column=3).value)
        
        if v1 is None: continue
        
        # Header Detection (Bold)
        if c1.font.bold and v3 is None:
            v1_up = v1.upper()
            mapped_name = None
            for key, val in FLEX_CATEGORY_WHITELIST.items():
                if key in v1_up:
                    mapped_name = val
                    break
            
            if mapped_name:
                existing = next((c for c in categories if c['name'] == mapped_name), None)
                if not existing:
                    curr_cat = {'name': mapped_name, 'subcategories': []}
                    categories.append(curr_cat)
                else:
                    curr_cat = existing
                
                curr_subcat = {'name': clean_text(v1), 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
                continue
            
            if curr_cat:
                if v3 is not None or isinstance(v2, (int, float)):
                    pass 
                else:
                    curr_subcat = {'name': clean_text(v1), 'products': []}
                    curr_cat['subcategories'].append(curr_subcat)
                    continue

        # Product Detection
        if v1 and (isinstance(v3, (int, float)) or isinstance(v2, (int, float))):
            if any(x in v1 for x in ['Julian', 'Jose Miguel', 'PRECIO', 'ESPESOR', 'rogamos']): continue
            if not curr_cat: continue 

            prod = {
                'name': clean_text(v1),
                'size': clean_text(v2) if not isinstance(v2, (int, float)) else '',
                'price': v3 if isinstance(v3, (int, float)) else v2
            }
            if not curr_subcat:
                curr_subcat = {'name': 'General', 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
            curr_subcat['products'].append(prod)

    for cat in categories:
        cat['subcategories'] = [sc for sc in cat['subcategories'] if sc['products']]
    return [c for c in categories if c['subcategories']]

def extract_rigidos():
    ws = wb['RIGIDOS']
    categories = []
    curr_cat = None
    curr_subcat = None
    
    NAMES = ['Julian', 'Jose Miguel']

    for row in range(1, ws.max_row + 1):
        # Read the whole row to be flexible
        row_vals = [strip_val(ws.cell(row=row, column=c).value) for c in range(1, 10)]
        row_fonts = [ws.cell(row=row, column=c).font.bold for c in range(1, 10)]
        v1, v2, v3 = row_vals[0], row_vals[1], row_vals[2]
        
        if not any(row_vals): continue
        
        # Category: Bold in Col 2, Col 1 is None
        if v2 and row_fonts[1] and not v1:
            if v2 in NAMES: continue
            curr_cat = {'name': clean_text(v2), 'subcategories': []}
            categories.append(curr_cat)
            curr_subcat = None
            continue
            
        # Subcategory: Bold in Col 1, Col 2 is None or metadata
        if v1 and row_fonts[0] and (not v2 or v2 == 'PRECIO' or v2 == 'MEDIDAS'):
            if v1 in NAMES: continue
            if not curr_cat:
                curr_cat = {'name': 'General', 'subcategories': []}
                categories.append(curr_cat)
            curr_subcat = {'name': clean_text(v1), 'products': []}
            curr_cat['subcategories'].append(curr_subcat)
            continue
            
        # Product Detection: Look for first numeric value in the row
        price = None
        price_half = None
        desc_parts = []
        for i, val in enumerate(row_vals):
            if isinstance(val, (int, float)) and price is None:
                price = val
                # Check next cell for price_half if common
                if i+1 < len(row_vals) and isinstance(row_vals[i+1], (int, float)):
                    price_half = row_vals[i+1]
                break
            elif isinstance(val, str) and price is None:
                if val.upper() not in ['PRECIO', 'ESPESOR', 'MEDIDAS', 'COLORES', '€ PLACA']:
                    desc_parts.append(val)
        
        if price is not None:
            full_name = " ".join(desc_parts)
            if any(n in full_name for n in NAMES): continue
            
            prod = {
                'name': clean_text(full_name),
                'price': price,
                'price_half': price_half
            }
            if not curr_subcat:
                if not curr_cat:
                    curr_cat = {'name': 'General', 'subcategories': []}
                    categories.append(curr_cat)
                curr_subcat = {'name': 'General', 'products': []}
                curr_cat['subcategories'].append(curr_subcat)
            curr_subcat['products'].append(prod)

    for cat in categories:
        cat['subcategories'] = [sc for sc in cat['subcategories'] if sc['products']]
    return [c for c in categories if c['subcategories']]

data = {
    'flexibles': extract_flexibles(),
    'rigidos': extract_rigidos()
}

with open('product_data_v5.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print('Extracted v5 data.')
